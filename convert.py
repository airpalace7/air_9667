#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
9667_테스트.xlsx -> data/*.json 변환 스크립트

사용법:
    python3 convert.py 9667_테스트.xlsx

엑셀을 수정한 뒤 이 스크립트를 다시 실행하면 data/ 폴더의 JSON 4개가
새로 생성됩니다. 그 JSON 파일들을 GitHub 저장소에 업로드(덮어쓰기)하면
웹페이지에 반영됩니다.
"""
import sys
import json
import re
import datetime
import openpyxl
import warnings
warnings.filterwarnings("ignore")

OUT_DIR = "data"


def hours_of(v):
    """timedelta -> 소수 시간(float). 아니면 None."""
    if isinstance(v, datetime.timedelta):
        return v.total_seconds() / 3600.0
    return None


def date_iso(v):
    """datetime -> 'YYYY-MM-DD'. 아니면 None."""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return None


TSN_BASE_RE = {
    "ac": re.compile(r"^=SUM\(G\d+-B\$3\)$"),
    "eng1": re.compile(r"^=SUM\(G\d+-\$?F\$?2\)$"),
    "eng2": re.compile(r"^=SUM\(G\d+-\$?H\$?2\)$"),
}


CELL_REF_RE = re.compile(r"^=([A-Z]+)(\d+)$")


def ref_info(ws, formula_val):
    """'=D41' 같은 단순 셀 참조 수식이면 참조 대상 행/항목명을 돌려준다. 아니면 None."""
    if not isinstance(formula_val, str):
        return None
    m = CELL_REF_RE.match(formula_val)
    if not m:
        return None
    ref_row = int(m.group(2))
    ref_item = fmt_val(ws.cell(ref_row, 1).value)[0]
    return {"row": ref_row, "item": ref_item}


def is_dash_or_none(v):
    return v is None or (isinstance(v, str) and v.strip() == "-")


def is_formula(v):
    return isinstance(v, str) and v.startswith("=")


def classify_row(ws_formula, r):
    """엑셀 수식을 보고 이 행이 표준 패턴(JS 재계산 가능)인지 판단."""
    f = ws_formula.cell(r, 6).value
    g = ws_formula.cell(r, 7).value
    h = ws_formula.cell(r, 8).value
    i = ws_formula.cell(r, 9).value

    ok_g = is_dash_or_none(g) or (isinstance(g, str) and re.match(r"^=SUM\(E\d+\+C\d+\)$", g))
    ok_f = is_dash_or_none(f) or (isinstance(f, str) and re.match(r"^=SUM\(D\d+\+B\d+\)$", f))
    ok_h = is_dash_or_none(h) or (isinstance(h, str) and re.match(r"^=F\d+-\$?L\$?2$", h))

    base = None
    for name, pat in TSN_BASE_RE.items():
        if isinstance(i, str) and pat.match(i):
            base = name
            break
    ok_i = is_dash_or_none(i) or (base is not None)

    recalculable = ok_f and ok_g and ok_h and ok_i
    return recalculable, base


def fmt_val(v):
    """셀 값을 표시용 문자열로 변환. 반환값: (문자열, 원값이 음수/초과인지 여부)"""
    if v is None:
        return "", False
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d"), False
    if isinstance(v, datetime.timedelta):
        total = v.total_seconds()
        neg = total < 0
        total = abs(total)
        total_hours = int(total // 3600)
        m = int((total % 3600) // 60)
        sign = "-" if neg else ""
        s = f"{sign}{total_hours}:{m:02d}"
        return s, neg
    if isinstance(v, (int, float)):
        neg = v < 0
        if isinstance(v, float) and not v.is_integer():
            s = f"{v:.2f}".rstrip("0").rstrip(".")
        else:
            s = str(int(v))
        return s, neg
    return str(v), False


def convert_insp(ws, ws_formula):
    info = {
        "등록기호": fmt_val(ws["B2"].value)[0],
        "TSN": fmt_val(ws["B3"].value)[0],
        "L/D": fmt_val(ws["D3"].value)[0],
        "1ENG_TSN": fmt_val(ws["F2"].value)[0],
        "2ENG_TSN": fmt_val(ws["H2"].value)[0],
        "1ENG_GG_CYC": fmt_val(ws["F3"].value)[0],
        "2ENG_GG_CYC": fmt_val(ws["H3"].value)[0],
        "1ENG_PT_CYC": fmt_val(ws["F4"].value)[0],
        "2ENG_PT_CYC": fmt_val(ws["H4"].value)[0],
        "DATE": fmt_val(ws["L2"].value)[0],
    }
    raw = {
        "ac_tsn_hours": hours_of(ws["B3"].value),
        "eng1_tsn_hours": hours_of(ws["F2"].value),
        "eng2_tsn_hours": hours_of(ws["H2"].value),
        "ld_cycle": ws["D3"].value if isinstance(ws["D3"].value, (int, float)) else None,
        "eng1_gg_cyc": ws["F3"].value if isinstance(ws["F3"].value, (int, float)) else None,
        "eng1_pt_cyc": ws["F4"].value if isinstance(ws["F4"].value, (int, float)) else None,
        "eng2_gg_cyc": ws["H3"].value if isinstance(ws["H3"].value, (int, float)) else None,
        "eng2_pt_cyc": ws["H4"].value if isinstance(ws["H4"].value, (int, float)) else None,
        "today": date_iso(ws["L2"].value),
    }
    sections = []
    cur = None
    # "제거일"/"재 장착일" 라벨 행을 만나면, 바로 앞서 추가한 실제 항목(예: CARGO HOOK 본체)에
    # 로그를 붙인다. 그 라벨 행 자체는 결과에 넣지 않는다(열 라벨일 뿐 별도 항목이 아님).
    last_row_obj = None
    last_row_num = None
    log_target = None  # 지금 로그를 채워넣고 있는 대상 row_obj (없으면 None)
    for r in range(7, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a and isinstance(a, str) and a.strip().startswith("▶"):
            cur = {"title": a.strip(), "rows": []}
            sections.append(cur)
            last_row_obj = None
            log_target = None
            continue
        if cur is None:
            continue
        if a is None and b is None:
            # item 칸이 비어있는 행. 로그 대상이 지정돼 있다면 이 행의 D/E열을
            # 제거일/재장착일 한 쌍으로 그 항목의 로그에 추가한다.
            if log_target is not None:
                removed = date_iso(ws.cell(r, 4).value)
                reinstalled = date_iso(ws.cell(r, 5).value)
                log_target["log_entries"].append({
                    "removed_date_iso": removed,
                    "reinstalled_date_iso": reinstalled,
                })
            continue

        performed_date_raw = ws.cell(r, 4).value
        performed_time_raw = ws.cell(r, 5).value
        # "제거일"/"재 장착일" 같은 열 라벨 행: 별도 항목이 아니라 바로 위 항목의
        # 로그 표 머리글일 뿐이므로, 결과에 추가하지 않고 앞 항목에 로그 기능을 붙인다.
        is_log_label_row = (
            isinstance(performed_date_raw, str) and "제거일" in performed_date_raw
            and isinstance(performed_time_raw, str) and "장착일" in performed_time_raw
        )
        if is_log_label_row:
            if last_row_obj is not None:
                last_row_obj["date_log"] = True
                last_row_obj["log_entries"] = []
                n_val = ws.cell(last_row_num, 14).value  # 숨겨진 N열: 실제 간격(일) 상수
                if isinstance(n_val, (int, float)):
                    last_row_obj["cargo_interval_days"] = int(n_val)
                last_row_obj["recalc"] = True  # 이제 웹에서 전용 로직으로 정확히 계산되므로
                log_target = last_row_obj
            continue

        item = fmt_val(a)[0]
        if not item:
            continue
        interval_day_raw = ws.cell(r, 2).value
        interval_time_raw = ws.cell(r, 3).value
        interval_day = fmt_val(interval_day_raw)[0]
        interval_time = fmt_val(interval_time_raw)[0]
        performed_date = fmt_val(performed_date_raw)[0]
        performed_time = fmt_val(performed_time_raw)[0]
        next_date = fmt_val(ws.cell(r, 6).value)[0]
        next_time = fmt_val(ws.cell(r, 7).value)[0]
        remain_day_raw = ws.cell(r, 8).value
        remain_day, neg_day = fmt_val(remain_day_raw)
        remain_time = fmt_val(ws.cell(r, 9).value)[0]
        remark = fmt_val(ws.cell(r, 10).value)[0]
        overdue = isinstance(remain_day_raw, (int, float)) and remain_day_raw < 0

        recalculable, tsn_base = classify_row(ws_formula, r)
        performed_date_ref = ref_info(ws, ws_formula.cell(r, 4).value)
        performed_time_ref = ref_info(ws, ws_formula.cell(r, 5).value)
        # 다른 행을 그대로 참조하는 단순 수식(예: =D41)일 때만 편집 불가로 잠근다.
        # 계산 상수로 들어간 수식(예: =(10938+6/60)/24)은 다른 셀을 참조하지 않으므로
        # 편집 가능하게 둔다 — 값 자체를 웹에서 바로 고칠 수 있어야 하기 때문.
        performed_date_editable = performed_date_ref is None
        performed_time_editable = performed_time_ref is None

        row_obj = {
            "item": item,
            "interval_day": interval_day,
            "interval_time": interval_time,
            "performed_date": performed_date,
            "performed_time": performed_time,
            "next_date": next_date,
            "next_time": next_time,
            "remain_day": remain_day,
            "remain_time": remain_time,
            "remark": remark,
            "overdue": overdue,
            "recalc": recalculable,
            "tsn_base": tsn_base,
            "performed_date_editable": performed_date_editable,
            "performed_time_editable": performed_time_editable,
            "performed_date_ref": performed_date_ref,
            "performed_time_ref": performed_time_ref,
            "interval_day_num": interval_day_raw if isinstance(interval_day_raw, (int, float)) else None,
            "interval_time_hours": hours_of(interval_time_raw),
            "performed_date_iso": date_iso(performed_date_raw),
            "performed_time_hours": hours_of(performed_time_raw),
        }
        cur["rows"].append(row_obj)
        last_row_obj = row_obj
        last_row_num = r
        log_target = None  # 새 일반 항목이 나왔으니 로그 채우기 대상은 해제(라벨 행 나오면 다시 지정)
    return {"info": info, "raw": raw, "sections": sections}


def classify_component_row(ws_formula, r):
    """부품/엔진 시트 한 행의 usage_time(13열)/next_exchange_time(15열)/remaining_time(17열)이
    표준 패턴(=SUM($C$기준행-H+K또는L), =SUM(H+F-K또는L), =SUM(O-$C$기준행))을 따르는지 확인한다.
    따르면 웹에서 값을 수정했을 때 실시간으로 재계산할 수 있고(recalc=True),
    아니면(참조 열이 서로 안 맞거나 아예 다른 형태면) 엑셀 계산값을 그대로 두고 편집을 막는다.
    """
    v13 = ws_formula.cell(r, 13).value
    v15 = ws_formula.cell(r, 15).value
    v17 = ws_formula.cell(r, 17).value

    m13 = re.match(r'^=SUM\(\$C\$(\d+)-H\d+\+([KL])\d+\)$', v13) if is_formula(v13) else None
    m15 = re.match(r'^=SUM\(H\d+\+F\d+-([KL])\d+\)$', v15) if is_formula(v15) else None
    m17 = re.match(r'^=SUM\(O\d+-\$C\$(\d+)\)$', v17) if is_formula(v17) else None

    has_interval = is_formula(v13) or is_formula(v15) or is_formula(v17)
    recalc = bool(m13 and m15 and m17 and m13.group(2) == m15.group(1) and m13.group(1) == m17.group(1))
    ref_row = int(m13.group(1)) if recalc else None
    baseline_col = m13.group(2) if recalc else None
    return recalc, has_interval, ref_row, baseline_col


def classify_date_row(ws_formula, r, sheet_kind):
    """AC TRP/ENG 시트의 "달력 일수" 기준(next_exchange_date=16열, remain=18열) 계산이
    표준 패턴을 따르는지 확인한다. sheet_kind: 'ac' 또는 'eng'."""
    v7 = ws_formula.cell(r, 7).value  # 교환주기(일수)
    v16 = ws_formula.cell(r, 16).value
    v18 = ws_formula.cell(r, 18).value

    interval_days = v7 if isinstance(v7, (int, float)) and v7 else None
    if interval_days is None:
        return {"date_recalc": False, "date_interval_days": None, "date_mode": None, "date_baseline_col": None}

    m18 = re.match(r'^=SUM\(P\d+-\$C\$1\)$', v18) if is_formula(v18) else None

    if sheet_kind == "ac":
        m16 = re.match(r'^=SUM\(I\d+\+G\d+-([KL])\d+\)$', v16) if is_formula(v16) else None
        date_recalc = bool(m16 and m18)
        baseline_col = m16.group(1) if m16 else None
    else:
        m16 = re.match(r'^=SUM\(([IL])\d+\+G\d+\)$', v16) if is_formula(v16) else None
        date_recalc = bool(m16 and m18)
        baseline_col = "L"

    return {
        "date_recalc": date_recalc,
        "date_interval_days": interval_days,
        "date_mode": sheet_kind if date_recalc else None,
        "date_baseline_col": baseline_col,
    }


def convert_component(ws, ws_formula, header_rows_end, data_start, ref_map, sheet_kind):
    """ref_map: {기준행번호: (insp.json raw의 키 이름, 'hours'|'cycles')}
    예: AC TRP는 {2: ('ac_tsn_hours','hours'), 3: ('ld_cycle','cycles')}
    sheet_kind: 'ac' 또는 'eng' - 달력 일수 기준 계산 방식이 시트마다 달라서 구분한다."""
    info_lines = []
    for r in range(1, header_rows_end):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if b:
            info_lines.append({"label": str(b), "value": fmt_val(c)[0]})
    rows = []
    for r in range(data_start, ws.max_row + 1):
        no = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if no is None and not name:
            continue
        remaining_time_raw = ws.cell(r, 17).value
        remaining_time, neg = fmt_val(remaining_time_raw)
        remaining_days_raw = ws.cell(r, 18).value
        overdue = isinstance(remaining_time_raw, (int, float)) and remaining_time_raw < 0
        if isinstance(remaining_time_raw, datetime.timedelta) and remaining_time_raw.total_seconds() < 0:
            overdue = True
        if isinstance(remaining_days_raw, (int, float)) and remaining_days_raw < 0:
            overdue = True
        next_date_raw = ws.cell(r, 16).value
        if isinstance(next_date_raw, datetime.datetime):
            today_cell = ws.parent["AC TRP, TBO"]["C1"].value if "AC TRP, TBO" in ws.parent.sheetnames else None
            today_dt = today_cell if isinstance(today_cell, datetime.datetime) else datetime.datetime.now()
            if next_date_raw < today_dt:
                overdue = True

        recalc, has_interval, ref_row, baseline_col = classify_component_row(ws_formula, r)
        ref_key, value_kind = (None, None)
        if ref_row is not None and ref_row in ref_map:
            ref_key, value_kind = ref_map[ref_row]
        else:
            recalc = False  # 알 수 없는 기준행이면 재계산하지 않음(엑셀관리)

        date_info = classify_date_row(ws_formula, r, sheet_kind)
        date_baseline_offset_days = None
        date_baseline_override_iso = None
        if date_info["date_recalc"]:
            bcol = date_info["date_baseline_col"]
            braw = ws.cell(r, 11 if bcol == "K" else 12).value
            if sheet_kind == "ac":
                # AC TRP: 보정 "일수"(대부분 0). 엑셀 SUM()은 텍스트를 0으로 취급하므로 그대로 따름.
                if isinstance(braw, (int, float)):
                    date_baseline_offset_days = int(braw)
                elif isinstance(braw, datetime.timedelta):
                    date_baseline_offset_days = int(braw.total_seconds() // 86400)
                else:
                    date_baseline_offset_days = 0
            else:
                # ENG: 보정 셀이 실제 날짜면 "기준 시작일"로 그 날짜를 쓰고(장착일 무시),
                # 아니면 장착일을 그대로 기준으로 쓴다.
                if isinstance(braw, datetime.datetime):
                    date_baseline_override_iso = date_iso(braw)

        installation_time_raw = ws.cell(r, 8).value
        exchange_cycle_raw = ws.cell(r, 6).value
        baseline_raw = ws.cell(r, 11 if baseline_col == "K" else 12).value if baseline_col else None

        def numeric_of(v):
            if isinstance(v, datetime.timedelta):
                return hours_of(v)
            if isinstance(v, (int, float)):
                return v
            return None

        # 참조 대상 다른 행의 값(=D41 같은 단순 참조) 여부 - 있으면 편집 불가
        name_ref = ref_info(ws, ws_formula.cell(r, 2).value)
        pn_ref = ref_info(ws, ws_formula.cell(r, 3).value)
        sn_ref = ref_info(ws, ws_formula.cell(r, 4).value)

        rows.append({
            "no": fmt_val(no)[0],
            "name": fmt_val(name)[0],
            "name_editable": name_ref is None and not is_formula(ws_formula.cell(r, 2).value),
            "pn": fmt_val(ws.cell(r, 3).value)[0],
            "pn_editable": pn_ref is None and not is_formula(ws_formula.cell(r, 3).value),
            "sn": fmt_val(ws.cell(r, 4).value)[0],
            "sn_editable": sn_ref is None and not is_formula(ws_formula.cell(r, 4).value),
            "type": fmt_val(ws.cell(r, 5).value)[0],
            "type_editable": not is_formula(ws_formula.cell(r, 5).value),
            "exchange_cycle": fmt_val(exchange_cycle_raw)[0],
            "exchange_cycle_num": numeric_of(exchange_cycle_raw),
            "exchange_cycle_editable": ref_info(ws, ws_formula.cell(r, 6).value) is None,
            "installation_time": fmt_val(installation_time_raw)[0],
            "installation_time_num": numeric_of(installation_time_raw),
            "installation_time_editable": ref_info(ws, ws_formula.cell(r, 8).value) is None,
            "installation_date": fmt_val(ws.cell(r, 9).value)[0],
            "installation_date_iso": date_iso(ws.cell(r, 9).value),
            "installation_date_editable": ref_info(ws, ws_formula.cell(r, 9).value) is None,
            "location": fmt_val(ws.cell(r, 10).value)[0],
            "location_editable": ref_info(ws, ws_formula.cell(r, 10).value) is None,
            "tsn": fmt_val(ws.cell(r, 11).value)[0],
            "usage_time": fmt_val(ws.cell(r, 13).value)[0],
            "next_exchange_time": fmt_val(ws.cell(r, 15).value)[0],
            "next_exchange_date": fmt_val(ws.cell(r, 16).value)[0],
            "next_exchange_date_iso": date_iso(ws.cell(r, 16).value),
            "remaining_time": remaining_time,
            "remaining_days": fmt_val(remaining_days_raw)[0],
            "note": fmt_val(ws.cell(r, 19).value)[0],
            "note_editable": ref_info(ws, ws_formula.cell(r, 19).value) is None,
            "overdue": overdue,
            "recalc": recalc,
            "has_interval": has_interval,
            "ref_key": ref_key,
            "value_kind": value_kind,
            "baseline_value": numeric_of(baseline_raw),
            "date_recalc": date_info["date_recalc"],
            "date_interval_days": date_info["date_interval_days"],
            "date_interval_days_editable": ref_info(ws, ws_formula.cell(r, 7).value) is None,
            "date_mode": date_info["date_mode"],
            "date_baseline_offset_days": date_baseline_offset_days,
            "date_baseline_override_iso": date_baseline_override_iso,
        })
    return {"info": info_lines, "rows": rows}


def main():
    if len(sys.argv) < 2:
        print("사용법: python3 convert.py <엑셀파일.xlsx>")
        sys.exit(1)
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, data_only=True)
    wb_formula = openpyxl.load_workbook(path, data_only=False)

    import os
    os.makedirs(OUT_DIR, exist_ok=True)

    data_insp = convert_insp(wb["HL9667 INSP"], wb_formula["HL9667 INSP"])
    with open(f"{OUT_DIR}/insp.json", "w", encoding="utf-8") as f:
        json.dump(data_insp, f, ensure_ascii=False, indent=2)

    data_trp = convert_component(
        wb["AC TRP, TBO"], wb_formula["AC TRP, TBO"], header_rows_end=9, data_start=10,
        ref_map={2: ("ac_tsn_hours", "hours"), 3: ("ld_cycle", "cycles")},
        sheet_kind="ac",
    )
    with open(f"{OUT_DIR}/trp.json", "w", encoding="utf-8") as f:
        json.dump(data_trp, f, ensure_ascii=False, indent=2)

    data_eng1 = convert_component(
        wb["#1 ENG TRP, TBO"], wb_formula["#1 ENG TRP, TBO"], header_rows_end=10, data_start=12,
        ref_map={6: ("eng1_tsn_hours", "hours"), 7: ("eng1_gg_cyc", "cycles"), 8: ("eng1_pt_cyc", "cycles")},
        sheet_kind="eng",
    )
    with open(f"{OUT_DIR}/eng1.json", "w", encoding="utf-8") as f:
        json.dump(data_eng1, f, ensure_ascii=False, indent=2)

    data_eng2 = convert_component(
        wb["#2 ENG TRP, TBO"], wb_formula["#2 ENG TRP, TBO"], header_rows_end=10, data_start=12,
        ref_map={6: ("eng2_tsn_hours", "hours"), 7: ("eng2_gg_cyc", "cycles"), 8: ("eng2_pt_cyc", "cycles")},
        sheet_kind="eng",
    )
    with open(f"{OUT_DIR}/eng2.json", "w", encoding="utf-8") as f:
        json.dump(data_eng2, f, ensure_ascii=False, indent=2)

    print("완료: data/insp.json, data/trp.json, data/eng1.json, data/eng2.json 생성됨")


if __name__ == "__main__":
    main()
